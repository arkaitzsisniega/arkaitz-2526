#!/usr/bin/env python3
"""
Importador de DISPAROS por partido desde `Goles TOTAL.xlsx`.

Origen: ~/Mi unidad/.../Estadisticas/Goles TOTAL.xlsx
Hoja: "RATIOS DISPAROS"

Una fila = un partido del Movistar Inter, con disparos a favor / en contra,
goles a favor / en contra, ratio (gol/disparo), minutos jugados.

Esta hoja sí tiene valores cacheados (Arkaitz la edita y guarda con valores).

Uso:
  /usr/bin/python3 src/estadisticas_disparos.py --upload
"""
from __future__ import annotations

import argparse
import datetime as _dt
import sys
import warnings
from pathlib import Path
from typing import Optional

warnings.filterwarnings("ignore")

import pandas as pd
def _col_letra(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


from openpyxl import load_workbook

XLSX_DEFAULT = (
    "/Users/mac/Mi unidad/Deporte/Futbol sala/Movistar Inter/"
    "2025-26/Estadisticas/Goles TOTAL.xlsx"
)
HOJA = "RATIOS DISPAROS"
HOJA_ZONAS = "ZONA GOLES"

CABECERAS = [
    "competicion", "rival", "fecha",
    "disparos_a_favor", "disparos_en_contra", "diferencia_disparos",
    "goles_a_favor", "goles_en_contra", "diferencia_goles",
    "ratio_a_favor", "ratio_en_contra",
    "minutos_jugados", "minutos_5x4_4x5",
    "disparos_af_1t",
]

# ZONA GOLES — mapa de columnas (1-indexed):
# A FAVOR:
#   Cuadrantes portería (P1-P9): cada cuadrante = 2 cols (disparos + goles).
#     BC=55, BD=56 → P1; BE-BF → P2; ...; BS-BT → P9 (cols 55-72)
#   Zonas campo (Z1-Z11): solo goles. BU=73 a CE=83 (11 cols).
# EN CONTRA:
#   Cuadrantes portería (P1-P9): CF=84, CG=85 → P1; ...; CV-CW → P9 (cols 84-101)
#   Zonas campo (Z1-Z11): solo goles. CX=102 a DH=112 (11 cols).
ZONA_COL_AF_P_INI = 55   # BC: D.AF P1
ZONA_COL_AF_Z_INI = 73   # BU: G.AF Z1
ZONA_COL_EC_P_INI = 84   # CF: D.EC P1
ZONA_COL_EC_Z_INI = 102  # CX: G.EC Z1


def _to_int(v) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def _to_int0(v) -> int:
    """Como _to_int pero devuelve 0 cuando no hay valor (para zonas)."""
    n = _to_int(v)
    return n if n is not None else 0


def _to_float(v) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, _dt.time):
        return v.hour * 60 + v.minute + v.second / 60
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _to_date_iso(v) -> str:
    if isinstance(v, _dt.datetime):
        return v.date().isoformat()
    if isinstance(v, _dt.date):
        return v.isoformat()
    return ""


def cargar_zonas(xlsx_path: str = XLSX_DEFAULT) -> pd.DataFrame:
    """Lee la hoja `ZONA GOLES` y extrae datos de cuadrantes portería
    (P1-P9: disparos + goles) y zonas campo (Z1-Z11: solo goles).

    Una fila por partido. Las cabeceras están en fila 2, los datos
    desde fila 3 (1-indexed). El bloque "ZONA GOLES" está en cols
    AZ (51) en adelante.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    if HOJA_ZONAS not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb[HOJA_ZONAS]
    filas: list[dict] = []
    for r in range(3, ws.max_row + 1):
        comp = ws.cell(r, 52).value   # col AZ (52)
        rival = ws.cell(r, 53).value  # col BA (53)
        fecha = ws.cell(r, 54).value  # col BB (54)
        if not comp or not rival:
            continue
        base = {
            "competicion": str(comp).strip(),
            "rival": str(rival).strip(),
            "fecha": _to_date_iso(fecha),
        }
        # A FAVOR · cuadrantes portería (cols BC..BT, 18 cols, 9 cuadrantes×2)
        for i in range(9):
            c_disp = ZONA_COL_AF_P_INI + 2 * i
            c_gol = c_disp + 1
            base[f"D_AF_P{i+1}"] = _to_int0(ws.cell(r, c_disp).value)
            base[f"G_AF_P{i+1}"] = _to_int0(ws.cell(r, c_gol).value)
        # A FAVOR · zonas campo (cols BU..CE, solo goles)
        for i in range(11):
            c = ZONA_COL_AF_Z_INI + i
            base[f"G_AF_Z{i+1}"] = _to_int0(ws.cell(r, c).value)
        # EN CONTRA · cuadrantes portería (cols CF..CW)
        for i in range(9):
            c_disp = ZONA_COL_EC_P_INI + 2 * i
            c_gol = c_disp + 1
            base[f"D_EC_P{i+1}"] = _to_int0(ws.cell(r, c_disp).value)
            base[f"G_EC_P{i+1}"] = _to_int0(ws.cell(r, c_gol).value)
        # EN CONTRA · zonas campo (cols CX..DH, solo goles)
        for i in range(11):
            c = ZONA_COL_EC_Z_INI + i
            base[f"G_EC_Z{i+1}"] = _to_int0(ws.cell(r, c).value)
        filas.append(base)
    return pd.DataFrame(filas)


def cargar(xlsx_path: str = XLSX_DEFAULT) -> pd.DataFrame:
    wb = load_workbook(xlsx_path, data_only=True)
    if HOJA not in wb.sheetnames:
        raise SystemExit(f"No encontré la hoja '{HOJA}' en {xlsx_path}")
    ws = wb[HOJA]
    filas: list[dict] = []
    # Datos a partir de fila 3 (1 = título, 2 = cabeceras)
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        comp = str(row[0]).strip() if row[0] else ""
        rival = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        if not comp or not rival:
            continue
        filas.append({
            "competicion": comp,
            "rival": rival,
            "fecha": _to_date_iso(row[2]) if len(row) > 2 else "",
            "disparos_a_favor":   _to_int(row[3])   if len(row) > 3 else None,
            "disparos_en_contra": _to_int(row[4])   if len(row) > 4 else None,
            "diferencia_disparos": _to_int(row[5])  if len(row) > 5 else None,
            "goles_a_favor":      _to_int(row[6])   if len(row) > 6 else None,
            "goles_en_contra":    _to_int(row[7])   if len(row) > 7 else None,
            "diferencia_goles":   _to_int(row[8])   if len(row) > 8 else None,
            "ratio_a_favor":      _to_float(row[9]) if len(row) > 9 else None,
            "ratio_en_contra":    _to_float(row[10]) if len(row) > 10 else None,
            "minutos_jugados":    _to_int(row[11])  if len(row) > 11 else None,
            "minutos_5x4_4x5":    _to_float(row[12]) if len(row) > 12 else None,
            "disparos_af_1t":     _to_int(row[13])  if len(row) > 13 else None,
        })
    return pd.DataFrame(filas, columns=CABECERAS)


def subir_a_sheet(df: pd.DataFrame, df_zonas: pd.DataFrame = None) -> None:
    import gspread
    from google.oauth2.service_account import Credentials
    SCOPES = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file("google_credentials.json", scopes=SCOPES)
    gc = gspread.authorize(creds)
    sh = gc.open("Arkaitz - Datos Temporada 2526")

    def _write(hoja: str, dataf: pd.DataFrame):
        try:
            ws = sh.worksheet(hoja)
            ws.clear()
        except gspread.exceptions.WorksheetNotFound:
            ws = sh.add_worksheet(title=hoja, rows=max(len(dataf) + 5, 100),
                                  cols=max(len(dataf.columns), 6))
        out = dataf.where(pd.notnull(dataf), "")
        valores = [list(out.columns)] + out.astype(str).values.tolist()
        ws.update(values=valores, range_name="A1")
        ws.format(f"A1:{_col_letra(len(out.columns))}1", {"textFormat": {"bold": True}})
        print(f"✅ {hoja}: {len(out)} filas, {len(out.columns)} cols")

    _write("EST_DISPAROS", df)
    if df_zonas is not None and not df_zonas.empty:
        _write("EST_DISPAROS_ZONAS", df_zonas)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=XLSX_DEFAULT)
    ap.add_argument("--upload", action="store_true")
    args = ap.parse_args()

    df = cargar(args.xlsx)
    print(f"Filas extraídas: {len(df)}")
    if not df.empty:
        print("Por competición:")
        print(df["competicion"].value_counts().to_string())
        print()
        print("Total disparos AF:", df["disparos_a_favor"].sum(skipna=True))
        print("Total disparos EC:", df["disparos_en_contra"].sum(skipna=True))
        print("Total goles AF:", df["goles_a_favor"].sum(skipna=True))
        print("Total goles EC:", df["goles_en_contra"].sum(skipna=True))

    df_zonas = cargar_zonas(args.xlsx)
    print(f"\nZONA GOLES: {len(df_zonas)} filas")
    if not df_zonas.empty:
        gaf = sum(int(df_zonas[f"G_AF_Z{i}"].sum()) for i in range(1, 12))
        gec = sum(int(df_zonas[f"G_EC_Z{i}"].sum()) for i in range(1, 12))
        print(f"  Total goles AF (zonas): {gaf}")
        print(f"  Total goles EC (zonas): {gec}")

    if args.upload:
        subir_a_sheet(df, df_zonas)


if __name__ == "__main__":
    main()
