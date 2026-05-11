#!/usr/bin/env python3
"""
Extractor de estadísticas de partido desde el Excel manual de Arkaitz.

Origen: ~/Mi unidad/Deporte/Futbol sala/Movistar Inter/2025-26/Estadisticas/Estadisticas2526.xlsx

Estrategia (ver docs/estadisticas_partidos.md):
- Solo leemos las hojas raw de cada partido (J*.RIVAL, AMIS.*, PLAYOFF*, C.*.*).
- Reimplementamos los agregados (EST.TOTAL, GOLES, TIEMPOS) en Python desde
  los eventos crudos. El Excel se edita en Numbers y sus fórmulas no
  están cacheadas; J.HERRERO sí lo está y nos sirve de validación.

Estructura fija dentro de cada hoja de partido:
  - Filas 5-19 (1-indexed) = 4-18 (0-indexed): rotaciones 1ª y 2ª parte
      cols B-K = 1ª-8ª rot 1T, col L = total 1T
      cols O-V = 1ª-8ª rot 2T, col W = total 2T
  - Filas 41-56 (40-55 0-indexed): eventos de goles
      col B (1) = marcador, D (3) = minuto, F (5) = acción,
      M (12) = portero, O,Q,S,U (14,16,18,20) = cuarteto,
      W (22) = goleador, Z (25) = asistente
"""
from __future__ import annotations

import argparse
import datetime as _dt
import re
import sys
import warnings
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Iterator, Optional

warnings.filterwarnings("ignore")

import pandas as pd
from openpyxl import load_workbook

from acciones import normalizar_accion

XLSX_DEFAULT = (
    "/Users/mac/Mi unidad/Deporte/Futbol sala/Movistar Inter/"
    "2025-26/Estadisticas/Estadisticas2526.xlsx"
)

# Patrones de hojas que SON partidos (con datos)
PATRON_PARTIDO = re.compile(
    r"^(?:J\d+\..+|AMIS\..+|AMISTOSO\..+|PLAYOFF\d+|SUP\..+|"
    r"C\.E\..+|C\.M\..+|C\.R\..+)$",
    re.IGNORECASE,
)

# Hojas que parecen partidos pero están vacías (plantillas no usadas)
HOJAS_VACIAS = {"J27", "J28", "J29", "J30", "P49", "CAJA NEGRA"}

# Porteros canónicos del equipo. Si en la columna "portero" del Excel
# aparece otro nombre, es un error de Arkaitz al rellenar (probablemente
# se confundió de columna). En ese caso movemos el nombre al cuarteto y
# dejamos el portero vacío (situación de portero-jugador o 5x4).
PORTEROS_CANONICOS = {"HERRERO", "GARCIA", "OSCAR"}

# Filas/columnas de los bloques (0-indexed)
ROT_FILA_INI, ROT_FILA_FIN = 4, 19          # 5..19 → idx 4..18, slice 4:19
ROT_COL_DORSAL, ROT_COL_NOMBRE = 1, 2       # B, C
# Rotaciones individuales: D..K (1ª-8ª rot 1T), O..V (1ª-8ª rot 2T).
# Col L y W son los totales calculados por Excel; los ignoramos para no
# duplicar al sumar.
ROT_COL_1T_INI, ROT_COL_1T_FIN = 3, 11      # D..K  (rangos slice)
ROT_COL_2T_INI, ROT_COL_2T_FIN = 14, 22     # O..V

EVT_FILA_INI, EVT_FILA_FIN = 41, 56         # 42..56 → idx 41..55, slice 41:56
EVT_COL_MARCADOR = 1   # B
EVT_COL_MIN = 3        # D
EVT_COL_ACCION = 5     # F
EVT_COL_PORTERO = 12   # M
EVT_COL_C1, EVT_COL_C2, EVT_COL_C3, EVT_COL_C4 = 14, 16, 18, 20  # O, Q, S, U
EVT_COL_GOLEADOR = 22  # W
EVT_COL_ASIST = 25     # Z

# Bloque de métricas individuales por jugador (filas 134-147)
# Cabecera en fila 133. Mapeo de columnas (0-indexed):
MET_FILA_INI, MET_FILA_FIN = 134, 148        # idx 133..147 → range 133..148 exclusive
MET_COL_DORSAL = 3       # D — Nº
MET_COL_JUGADOR = 4      # E — JUGADOR
MET_COL_MINS = 6         # G — MINS (timedelta)
MET_COL_PF = 7           # H — Pérdidas Forzadas
MET_COL_PNF = 8          # I — Pérdidas No Forzadas
MET_COL_ROBOS = 9        # J
MET_COL_CORTES = 10      # K
MET_COL_BDG = 12         # M — Balón Dividido Ganado
MET_COL_BDP = 13         # N — Balón Dividido Perdido
MET_COL_DP = 15          # P — Disparos a Puerta
MET_COL_DPALO = 16       # Q — Disparos al Palo
MET_COL_DB = 17          # R — Disparos Bloqueados
MET_COL_DF = 18          # S — Disparos Fuera
MET_COL_OUT = 19         # T — OUT (campo)
# Columnas de PORTERO (solo aplica a porteros):
MET_COL_POSTE = 20       # U
MET_COL_BLOQ = 21        # V — bloqueos del portero
MET_COL_PAR = 22         # W — paradas
MET_COL_GOL_PORT = 23    # X — goles encajados
MET_COL_TA = 24          # Y — Tarjeta Amarilla
MET_COL_TR = 25          # Z — Tarjeta Roja


def clasificar_competicion(nombre_hoja: str) -> tuple[str, str]:
    """Devuelve (tipo, competicion_legible)."""
    n = nombre_hoja.upper()
    if n.startswith("J") and "." in n and re.match(r"^J\d+\.", n):
        return ("LIGA", "Liga 25/26")
    if n.startswith("AMISTOSO."):
        return ("AMISTOSO", "Amistoso (temporada)")
    if n.startswith("AMIS."):
        return ("AMISTOSO", "Amistoso pretemporada")
    if n.startswith("PLAYOFF"):
        return ("PLAYOFF", "Playoff Liga")
    if n.startswith("SUP."):
        return ("SUPERCOPA", "Supercopa")
    if n.startswith("C.E."):
        return ("COPA_ESPANA", "Copa de España")
    if n.startswith("C.M."):
        return ("COPA_MUNDO", "Copa del Mundo de Clubes")
    if n.startswith("C.R."):
        return ("COPA_REY", "Copa del Rey")
    return ("OTRO", "Otro")


def extraer_rival(nombre_hoja: str) -> str:
    """Saca el rival del nombre de la pestaña."""
    n = nombre_hoja
    # Quitar prefijo Jx., AMIS., etc.
    n = re.sub(r"^J\d+\.\s*", "", n)
    n = re.sub(r"^AMIS\.\s*", "", n)
    n = re.sub(r"^AMISTOSO\.\s*", "", n)
    n = re.sub(r"^C\.[EMR]\.(?:\d+ª\.)?\s*", "", n)
    n = re.sub(r"^SUP\.\s*", "", n)
    return n.strip().upper()


def _to_minutes(v) -> float:
    """Convierte una celda con duración a minutos float.
    Acepta: time, timedelta, datetime, str 'H:MM:SS', float (días)."""
    if v is None or v == "" or (isinstance(v, str) and not v.strip()):
        return 0.0
    if isinstance(v, _dt.time):
        return v.hour * 60 + v.minute + v.second / 60.0
    if isinstance(v, _dt.timedelta):
        return v.total_seconds() / 60.0
    if isinstance(v, _dt.datetime):
        return v.hour * 60 + v.minute + v.second / 60.0
    if isinstance(v, (int, float)):
        # Excel almacena duraciones como fracción de día
        if 0 <= v <= 1:
            return v * 24 * 60
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        # Acepta ":2:57" (typo de Arkaitz, sin la hora)
        if s.startswith(":"):
            s = "0" + s
        m = re.match(r"^(\d+):(\d{1,2}):(\d{1,2})$", s)
        if m:
            h, mi, se = map(int, m.groups())
            return h * 60 + mi + se / 60.0
        m = re.match(r"^(\d+):(\d{1,2})$", s)
        if m:
            mi, se = map(int, m.groups())
            return mi + se / 60.0
    return 0.0


def _to_minute_int(v) -> Optional[int]:
    """Convierte el campo MIN del evento a minuto entero (1..40)."""
    m = _to_minutes(v)
    if m <= 0:
        return None
    return int(round(m))


def _to_minute_float(v) -> Optional[float]:
    """Convierte el campo MIN del evento a minutos como float (mm.ss/60),
    p.ej. 12:37 -> 12.6166. Permite formato mm:ss en el PDF."""
    m = _to_minutes(v)
    if m <= 0:
        return None
    return round(m, 4)


def _norm_nombre(s) -> str:
    if s is None:
        return ""
    return str(s).strip().upper()


@dataclass
class JugadorEnPartido:
    partido_id: str
    tipo: str
    competicion: str
    rival: str
    fecha: Optional[_dt.date]
    dorsal: Optional[int]
    jugador: str
    min_1t: float
    min_2t: float
    min_total: float
    convocado: bool
    participa: bool
    # Rotaciones individuales (1ª-8ª de cada parte). En minutos float.
    # Los valores no rellenados quedan en 0.0.
    rot_1t_1: float = 0.0
    rot_1t_2: float = 0.0
    rot_1t_3: float = 0.0
    rot_1t_4: float = 0.0
    rot_1t_5: float = 0.0
    rot_1t_6: float = 0.0
    rot_1t_7: float = 0.0
    rot_1t_8: float = 0.0
    rot_2t_1: float = 0.0
    rot_2t_2: float = 0.0
    rot_2t_3: float = 0.0
    rot_2t_4: float = 0.0
    rot_2t_5: float = 0.0
    rot_2t_6: float = 0.0
    rot_2t_7: float = 0.0
    rot_2t_8: float = 0.0
    # Métricas individuales de juego (de filas 134-147 en cada partido)
    pf: int = 0           # Pérdidas Forzadas
    pnf: int = 0          # Pérdidas No Forzadas
    robos: int = 0
    cortes: int = 0
    bdg: int = 0          # Balón Dividido Ganado
    bdp: int = 0          # Balón Dividido Perdido
    dp: int = 0           # Disparos a Puerta
    dpalo: int = 0        # Disparos al Palo
    db: int = 0           # Disparos Bloqueados
    df: int = 0           # Disparos Fuera
    out: int = 0          # OUT (jugadores de campo)
    # Columnas de portero (vacías para jugadores de campo)
    poste_p: int = 0      # Tiros del rival al poste (portero)
    bloq_p: int = 0       # Bloqueos del portero
    par: int = 0          # Paradas del portero
    gol_p: int = 0        # Goles encajados estando el portero en pista
    salida: int = 0       # Salida CORRECTA del portero (nuevo iter 3)
    salida_fallida: int = 0  # Salida fallida del portero (nuevo iter 3)
    ta: int = 0           # Tarjeta Amarilla
    tr: int = 0           # Tarjeta Roja (nuevo iter 3)


@dataclass
class TotalesPartido:
    """Datos a nivel partido (no por jugador): disparos totales del Inter
    y del rival, calculados de la fila 150 del Excel y de la suma de
    métricas de los porteros."""
    partido_id: str
    tipo: str
    competicion: str
    rival: str
    fecha: Optional[_dt.date]
    # Cabecera (filas 2-3 del Excel: PARTIDO/CATEGORÍA/LUGAR/HORA/FECHA)
    categoria: str = ""    # ej. "LIGA 25/26", "COPA REY"
    lugar: str = ""        # ej. "MADRID", "BARCELONA", "CARTAGENA"
    hora: str = ""         # ej. "13:00h"
    local_visitante: str = ""  # "LOCAL" / "VISITANTE" (deducido por lugar)
    # Inter (calculados desde fila 149)
    dp_inter: int = 0
    dpalo_inter: int = 0
    db_inter: int = 0
    df_inter: int = 0
    dt_inter: int = 0     # disparos totales Inter (fila 150 col P)
    pf_inter: int = 0
    pnf_inter: int = 0
    robos_inter: int = 0
    cortes_inter: int = 0
    # Rival (calculados desde fila 150 col T y desde porteros)
    dt_rival: int = 0     # disparos totales rival (fila 150 col T)
    dp_rival: int = 0     # paradas + goles encajados (suma de porteros)
    dpalo_rival: int = 0  # poste recibido (suma de porteros)
    db_rival: int = 0     # bloqueos del portero
    # Goles del partido
    goles_a_favor: int = 0
    goles_en_contra: int = 0


@dataclass
class EventoGol:
    partido_id: str
    tipo: str
    competicion: str
    rival: str
    fecha: Optional[_dt.date]
    minuto: Optional[int]
    intervalo_5min: str
    accion_raw: str         # texto bruto del Excel (ej. "AF.4x4")
    accion: str             # canónico (ej. "Ataque Posicional 4x4")
    marcador: str
    equipo_marca: str
    goleador: str
    asistente: str
    portero: str
    cuarteto: list[str]
    descripcion: str = ""   # texto libre que Arkaitz escribe a mano en el Sheet
                            # (se preserva entre re-extracciones)
    minuto_mmss: str = ""   # minuto en formato mm:ss (ej. "12:37"). Si no
                            # disponible, fallback a {minuto:02d}:00 al render.


def _intervalo_5min(minuto: Optional[int]) -> str:
    if not minuto:
        return ""
    if minuto < 1:
        return ""
    # 1..5 → "0-5", 6..10 → "5-10", etc.
    base = ((minuto - 1) // 5) * 5
    return f"{base}-{base + 5}"


def parsear_fecha_hora(ws_values) -> Optional[_dt.date]:
    """Busca una fecha en filas 1-3, cols X-Y aprox."""
    for r_idx in range(0, 4):
        if r_idx >= len(ws_values):
            continue
        row = ws_values[r_idx]
        for c_idx in range(15, min(len(row), 30)):
            v = row[c_idx]
            if isinstance(v, (_dt.date, _dt.datetime)):
                return v.date() if isinstance(v, _dt.datetime) else v
    return None


def parsear_partido(
    nombre_hoja: str, valores: list[list]
) -> tuple[list[JugadorEnPartido], list[EventoGol]]:
    """Devuelve (jugadores_en_partido, eventos_de_gol)."""
    tipo, competicion = clasificar_competicion(nombre_hoja)
    rival = extraer_rival(nombre_hoja)
    fecha = parsear_fecha_hora(valores)

    jugadores: list[JugadorEnPartido] = []
    eventos: list[EventoGol] = []

    # ─── Rotaciones (filas 5-19, idx 4-18) ─────────────────────────────────
    for r_idx in range(ROT_FILA_INI, min(ROT_FILA_FIN, len(valores))):
        row = valores[r_idx]
        if not row or len(row) <= ROT_COL_NOMBRE:
            continue
        nombre = _norm_nombre(row[ROT_COL_NOMBRE])
        if not nombre or nombre in ("NOMBRE", "JUGADOR"):
            continue
        try:
            dorsal = int(row[ROT_COL_DORSAL]) if row[ROT_COL_DORSAL] is not None else None
        except (TypeError, ValueError):
            dorsal = None

        # Rotaciones individuales 1T (cols D..K, idx 3..10)
        rots_1t = []
        for c in range(ROT_COL_1T_INI, ROT_COL_1T_FIN):
            v = _to_minutes(row[c]) if c < len(row) else 0.0
            rots_1t.append(round(v, 2))
        # Pad a 8
        rots_1t = (rots_1t + [0.0] * 8)[:8]
        min_1t = sum(rots_1t)

        # Rotaciones individuales 2T (cols O..V, idx 14..21)
        rots_2t = []
        for c in range(ROT_COL_2T_INI, ROT_COL_2T_FIN):
            v = _to_minutes(row[c]) if c < len(row) else 0.0
            rots_2t.append(round(v, 2))
        rots_2t = (rots_2t + [0.0] * 8)[:8]
        min_2t = sum(rots_2t)

        min_total = min_1t + min_2t
        convocado = True
        participa = min_total > 0

        jp = JugadorEnPartido(
            partido_id=nombre_hoja,
            tipo=tipo,
            competicion=competicion,
            rival=rival,
            fecha=fecha,
            dorsal=dorsal,
            jugador=nombre,
            min_1t=round(min_1t, 2),
            min_2t=round(min_2t, 2),
            min_total=round(min_total, 2),
            convocado=convocado,
            participa=participa,
        )
        # Inyectar rotaciones individuales
        for i in range(8):
            setattr(jp, f"rot_1t_{i+1}", rots_1t[i])
            setattr(jp, f"rot_2t_{i+1}", rots_2t[i])
        jugadores.append(jp)

    # ─── Eventos de gol (filas 41-56, idx 40-55) ───────────────────────────
    for r_idx in range(EVT_FILA_INI - 1, min(EVT_FILA_FIN, len(valores))):
        row = valores[r_idx]
        if not row or len(row) <= EVT_COL_GOLEADOR:
            continue
        goleador = _norm_nombre(row[EVT_COL_GOLEADOR])
        accion = _norm_nombre(row[EVT_COL_ACCION]) if EVT_COL_ACCION < len(row) else ""
        minuto = _to_minute_int(row[EVT_COL_MIN]) if EVT_COL_MIN < len(row) else None
        # Versión float (para mm:ss). Lo guardamos como string "MM:SS".
        minuto_f = _to_minute_float(row[EVT_COL_MIN]) if EVT_COL_MIN < len(row) else None
        if minuto_f is not None and minuto_f > 0:
            mm = int(minuto_f)
            ss = int(round((minuto_f - mm) * 60))
            if ss == 60:
                mm += 1; ss = 0
            minuto_mmss = f"{mm:02d}:{ss:02d}"
        else:
            minuto_mmss = ""
        if not goleador and not accion and not minuto:
            continue
        if goleador in ("GOLEADOR", "RIVAL/JUGADOR"):  # cabeceras residuales
            continue

        marcador = ""
        if EVT_COL_MARCADOR < len(row) and row[EVT_COL_MARCADOR] is not None:
            marcador = str(row[EVT_COL_MARCADOR]).strip()

        equipo_marca = "RIVAL" if goleador == "RIVAL" else "INTER"
        portero = _norm_nombre(row[EVT_COL_PORTERO]) if EVT_COL_PORTERO < len(row) else ""
        cuarteto = []
        for c in (EVT_COL_C1, EVT_COL_C2, EVT_COL_C3, EVT_COL_C4):
            if c < len(row):
                v = _norm_nombre(row[c])
                if v:
                    cuarteto.append(v)

        # Validación de portero canónico: si lo que viene en col M no es
        # uno de los 3 porteros oficiales, es error de apuntado. En ese
        # caso lo movemos al cuarteto (era jugador de campo) y el portero
        # queda vacío (situación de 5 jugadores en pista, portero-jugador).
        if portero and portero not in PORTEROS_CANONICOS:
            if portero not in cuarteto:
                cuarteto.append(portero)
            portero = ""

        asist = ""
        if EVT_COL_ASIST < len(row):
            asist = _norm_nombre(row[EVT_COL_ASIST])

        accion_canon = normalizar_accion(accion, equipo_marca)
        eventos.append(EventoGol(
            partido_id=nombre_hoja,
            tipo=tipo,
            competicion=competicion,
            rival=rival,
            fecha=fecha,
            minuto=minuto,
            intervalo_5min=_intervalo_5min(minuto),
            accion_raw=accion,
            accion=accion_canon,
            marcador=marcador,
            equipo_marca=equipo_marca,
            goleador=goleador,
            asistente=asist,
            portero=portero,
            cuarteto=cuarteto,
            descripcion="",
            minuto_mmss=minuto_mmss,
        ))

    # ─── Métricas individuales (filas 134-147) ──────────────────────────────
    # Construimos un dict por nombre de jugador para inyectar luego en jugadores.
    metricas_por_jugador: dict[str, dict] = {}
    for r_idx in range(MET_FILA_INI - 1, min(MET_FILA_FIN, len(valores))):
        row = valores[r_idx]
        if not row or len(row) <= MET_COL_JUGADOR:
            continue
        nombre = _norm_nombre(row[MET_COL_JUGADOR])
        if not nombre or nombre in ("EQUIPO", "TOTALES", "JUGADOR"):
            continue
        # Minutos totales del bloque MET (col G, timedelta normalmente).
        # Lo usamos como FALLBACK si las rotaciones (filas 5-19) están vacías,
        # y para añadir partidos donde Arkaitz no rellena las rotaciones detalladas.
        mins_met = 0.0
        if MET_COL_MINS < len(row):
            mins_met = _to_minutes(row[MET_COL_MINS])

        metricas_por_jugador[nombre] = {
            "pf":      _to_int(row, MET_COL_PF),
            "pnf":     _to_int(row, MET_COL_PNF),
            "robos":   _to_int(row, MET_COL_ROBOS),
            "cortes":  _to_int(row, MET_COL_CORTES),
            "bdg":     _to_int(row, MET_COL_BDG),
            "bdp":     _to_int(row, MET_COL_BDP),
            "dp":      _to_int(row, MET_COL_DP),
            "dpalo":   _to_int(row, MET_COL_DPALO),
            "db":      _to_int(row, MET_COL_DB),
            "df":      _to_int(row, MET_COL_DF),
            "out":     _to_int(row, MET_COL_OUT),
            "poste_p": _to_int(row, MET_COL_POSTE),
            "bloq_p":  _to_int(row, MET_COL_BLOQ),
            "par":     _to_int(row, MET_COL_PAR),
            "gol_p":   _to_int(row, MET_COL_GOL_PORT),
            "ta":      _to_int(row, MET_COL_TA),
            "tr":      _to_int(row, MET_COL_TR),
            "_mins_met": round(mins_met, 2),  # auxiliar, no se sube al Sheet
        }

    # Inyectar en los JugadorEnPartido ya creados
    for j in jugadores:
        m = metricas_por_jugador.get(j.jugador)
        if m:
            mins_met = m.pop("_mins_met", 0.0)
            for k, v in m.items():
                setattr(j, k, v)
            # Si la sección de rotaciones está vacía pero el bloque MET tiene
            # minutos, usar éstos como fuente principal. Las rotaciones
            # individuales quedarán a 0 (no es un error, simplemente Arkaitz
            # no las rellenó este partido).
            if j.min_total == 0 and mins_met > 0:
                j.min_total = mins_met
                j.participa = True

    # Añadir jugadores que están SOLO en el bloque de métricas pero no en la
    # tabla de rotaciones (pasa cuando Arkaitz no rellenó las rotaciones).
    nombres_existentes = {j.jugador for j in jugadores}
    for nombre, m in metricas_por_jugador.items():
        if nombre in nombres_existentes:
            continue
        mins_met = m.pop("_mins_met", 0.0)
        if mins_met <= 0:
            continue
        jp = JugadorEnPartido(
            partido_id=nombre_hoja, tipo=tipo, competicion=competicion,
            rival=rival, fecha=fecha, dorsal=None, jugador=nombre,
            min_1t=0.0, min_2t=0.0, min_total=mins_met,
            convocado=True, participa=True,
        )
        for k, v in m.items():
            setattr(jp, k, v)
        jugadores.append(jp)

    # ─── Totales del partido (fila 149-150) ─────────────────────────────────
    totales = TotalesPartido(
        partido_id=nombre_hoja,
        tipo=tipo,
        competicion=competicion,
        rival=rival,
        fecha=fecha,
    )

    # ─── Cabecera del partido (filas 2-3 del Excel) ─────────────────────────
    # E2/L2/P2/T2/W2 = labels  → E3/I3/L3/P3/T3/W3 = valores
    # E=col 4 (índice 4), I=8, L=11, P=15, T=19, W=22
    if len(valores) >= 3:
        row3 = valores[2]
        def _cell(idx):
            if idx >= len(row3):
                return ""
            v = row3[idx]
            if v is None or v == "":
                return ""
            if isinstance(v, _dt.time):
                return v.strftime("%H:%M") + "h"
            if isinstance(v, _dt.datetime):
                return v.strftime("%d/%m/%Y")
            return str(v).strip()
        totales.categoria = _cell(11)   # L3
        totales.lugar     = _cell(15)   # P3
        totales.hora      = _cell(19)   # T3
        # Local/Visitante: el Movistar Inter juega como local en pabellones
        # de Madrid (Garbajosa, Magariños, Madrid). Cualquier otro lugar es
        # visitante. Si lugar está vacío, dejar el campo vacío.
        lugar_up = totales.lugar.upper()
        es_local = any(k in lugar_up for k in
                       ("MADRID", "MAGARI", "GARBAJOSA", "TORREJON",
                        "ALCALA", "ALCOBENDAS"))
        if es_local:
            totales.local_visitante = "LOCAL"
        elif lugar_up:
            totales.local_visitante = "VISITANTE"

    # Fila 149: TOTALES de Inter
    if len(valores) > 148:
        row_t = valores[148]
        if row_t and len(row_t) > MET_COL_DF:
            totales.pf_inter     = _to_int(row_t, MET_COL_PF)
            totales.pnf_inter    = _to_int(row_t, MET_COL_PNF)
            totales.robos_inter  = _to_int(row_t, MET_COL_ROBOS)
            totales.cortes_inter = _to_int(row_t, MET_COL_CORTES)
            totales.dp_inter     = _to_int(row_t, MET_COL_DP)
            totales.dpalo_inter  = _to_int(row_t, MET_COL_DPALO)
            totales.db_inter     = _to_int(row_t, MET_COL_DB)
            totales.df_inter     = _to_int(row_t, MET_COL_DF)
    # Fila 150: "DISPAROS TOTALES" del partido
    if len(valores) > 149:
        row_dt = valores[149]
        if row_dt and len(row_dt) > 19:
            totales.dt_inter = _to_int(row_dt, 15)  # col P
            totales.dt_rival = _to_int(row_dt, 19)  # col T
    # Disparos del rival a porterías (sumando paradas + goles encajados de los porteros)
    porteros = [j for j in jugadores if j.par > 0 or j.gol_p > 0 or j.bloq_p > 0 or j.poste_p > 0]
    for p in porteros:
        totales.dp_rival    += p.par + p.gol_p
        totales.dpalo_rival += p.poste_p
        totales.db_rival    += p.bloq_p
    # Goles
    totales.goles_a_favor   = sum(1 for e in eventos if e.equipo_marca == "INTER")
    totales.goles_en_contra = sum(1 for e in eventos if e.equipo_marca == "RIVAL")

    return jugadores, eventos, totales


def _to_int(row, col: int) -> int:
    """Lee row[col] como int seguro. Devuelve 0 si vacío, None, o no parseable."""
    if col >= len(row):
        return 0
    v = row[col]
    if v is None or v == "":
        return 0
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)):
        try:
            return int(v)
        except (ValueError, OverflowError):
            return 0
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return 0
        try:
            return int(float(s.replace(",", ".")))
        except ValueError:
            return 0
    return 0


def hoja_es_partido(nombre: str) -> bool:
    if nombre in HOJAS_VACIAS:
        return False
    return bool(PATRON_PARTIDO.match(nombre))


def cargar_valores_hoja(wb, nombre: str) -> list[list]:
    """Devuelve la hoja como lista de listas, valores Python."""
    ws = wb[nombre]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def procesar_excel(xlsx_path: str = XLSX_DEFAULT) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    wb = load_workbook(xlsx_path, data_only=True)
    todos_jug: list[dict] = []
    todos_evt: list[dict] = []
    todos_tot: list[dict] = []

    for nombre in wb.sheetnames:
        if not hoja_es_partido(nombre):
            continue
        valores = cargar_valores_hoja(wb, nombre)
        jugadores, eventos, totales = parsear_partido(nombre, valores)
        if not jugadores:
            continue
        if not any(j.participa for j in jugadores):
            continue
        for j in jugadores:
            todos_jug.append(asdict(j))
        for e in eventos:
            d = asdict(e)
            d["cuarteto"] = "|".join(d["cuarteto"])
            todos_evt.append(d)
        todos_tot.append(asdict(totales))

    df_jug = pd.DataFrame(todos_jug)
    df_evt = pd.DataFrame(todos_evt)
    df_tot = pd.DataFrame(todos_tot)

    # Cruzar: goles a favor, en contra, asistencias por (partido_id, jugador)
    if not df_evt.empty and not df_jug.empty:
        # Goles: solo los de INTER cuentan como "goles_a_favor" del goleador
        gf = df_evt[df_evt["equipo_marca"] == "INTER"].groupby(
            ["partido_id", "goleador"]).size().reset_index(name="goles_a_favor")
        gf = gf.rename(columns={"goleador": "jugador"})
        df_jug = df_jug.merge(gf, on=["partido_id", "jugador"], how="left")
        df_jug["goles_a_favor"] = df_jug["goles_a_favor"].fillna(0).astype(int)

        # Asistencias: las que aparecen en col Z
        ast = df_evt[df_evt["asistente"].astype(bool)].groupby(
            ["partido_id", "asistente"]).size().reset_index(name="asistencias")
        ast = ast.rename(columns={"asistente": "jugador"})
        df_jug = df_jug.merge(ast, on=["partido_id", "jugador"], how="left")
        df_jug["asistencias"] = df_jug["asistencias"].fillna(0).astype(int)
    else:
        df_jug["goles_a_favor"] = 0
        df_jug["asistencias"] = 0

    return df_jug, df_evt, df_tot


def calcular_agregados_jugador(df_jug: pd.DataFrame) -> pd.DataFrame:
    """Construye _VISTA_EST_JUGADOR: una fila por jugador con totales.

    Incluye también desgloses por competición (LIGA, COPA_*, AMISTOSO).
    """
    if df_jug.empty:
        return pd.DataFrame()

    # Total por jugador
    g = df_jug.groupby("jugador", as_index=False).agg(
        partidos_convocado=("convocado", "sum"),
        partidos_jugados=("participa", "sum"),
        min_total=("min_total", "sum"),
        min_1t=("min_1t", "sum"),
        min_2t=("min_2t", "sum"),
        goles=("goles_a_favor", "sum"),
        asistencias=("asistencias", "sum"),
    )
    g["min_por_partido"] = (g["min_total"] / g["partidos_jugados"].clip(lower=1)).round(2)
    g["gol_y_asist"] = g["goles"] + g["asistencias"]

    # Pivot por competición (solo minutos y goles para no inflar)
    pivot_min = (df_jug.groupby(["jugador", "tipo"])["min_total"].sum()
                 .unstack(fill_value=0))
    pivot_min.columns = [f"min_{c.lower()}" for c in pivot_min.columns]
    pivot_gol = (df_jug.groupby(["jugador", "tipo"])["goles_a_favor"].sum()
                 .unstack(fill_value=0))
    pivot_gol.columns = [f"goles_{c.lower()}" for c in pivot_gol.columns]
    g = g.merge(pivot_min, on="jugador", how="left").merge(pivot_gol, on="jugador", how="left")
    g = g.fillna(0)
    return g.sort_values("goles", ascending=False)


def _col_letra(n: int) -> str:
    """1 → A, 2 → B, ..., 27 → AA, 28 → AB, ..."""
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def subir_a_sheet(df_jug: pd.DataFrame, df_evt: pd.DataFrame, df_agr: pd.DataFrame,
                  df_tot: pd.DataFrame = None) -> None:
    """Sube las 3 hojas al Sheet maestro (mismo Sheet que el resto del proyecto)."""
    import gspread
    from google.oauth2.service_account import Credentials
    SCOPES = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    SHEET_NAME = "Arkaitz - Datos Temporada 2526"
    creds = Credentials.from_service_account_file("google_credentials.json", scopes=SCOPES)
    gc = gspread.authorize(creds)
    sh = gc.open(SHEET_NAME)

    def _write(hoja: str, df: pd.DataFrame):
        # Convertir fechas a string ISO para evitar serializaciones raras
        out = df.copy()
        for col in out.columns:
            if pd.api.types.is_datetime64_any_dtype(out[col]):
                out[col] = out[col].dt.strftime("%Y-%m-%d").fillna("")
            elif out[col].dtype == "object":
                out[col] = out[col].apply(
                    lambda v: v.strftime("%Y-%m-%d") if isinstance(v, _dt.date) and not isinstance(v, _dt.datetime)
                    else (v.strftime("%Y-%m-%d") if isinstance(v, _dt.datetime) else v)
                )
        # Reemplazar NaN/None por ""
        out = out.where(pd.notnull(out), "")
        # Convertir bool a int (Sheets los guarda mejor)
        for col in out.columns:
            if out[col].dtype == bool:
                out[col] = out[col].astype(int)

        try:
            ws = sh.worksheet(hoja)
            ws.clear()
        except gspread.exceptions.WorksheetNotFound:
            ws = sh.add_worksheet(title=hoja, rows=max(len(out) + 10, 100), cols=max(len(out.columns), 6))

        valores = [list(out.columns)] + out.astype(str).values.tolist()
        ws.update(values=valores, range_name="A1")
        ws.format(f"A1:{_col_letra(len(out.columns))}1", {"textFormat": {"bold": True}})
        print(f"  ✅ {hoja}: {len(out)} filas, {len(out.columns)} cols")

    # ─── Preservar descripciones de gol escritas a mano ─────────────────────
    if not df_evt.empty:
        try:
            ws_old = sh.worksheet("EST_EVENTOS")
            old_rows = ws_old.get_all_records()
            old_df = pd.DataFrame(old_rows)
            if not old_df.empty and "descripcion" in old_df.columns:
                # Clave única: (partido_id, minuto, goleador, equipo_marca)
                old_df["_k"] = (
                    old_df["partido_id"].astype(str) + "|" +
                    old_df["minuto"].astype(str) + "|" +
                    old_df["goleador"].astype(str) + "|" +
                    old_df["equipo_marca"].astype(str)
                )
                desc_map = dict(zip(old_df["_k"], old_df["descripcion"].astype(str)))
                df_evt["_k"] = (
                    df_evt["partido_id"].astype(str) + "|" +
                    df_evt["minuto"].astype(str) + "|" +
                    df_evt["goleador"].astype(str) + "|" +
                    df_evt["equipo_marca"].astype(str)
                )
                # Solo sobreescribimos si el viejo tenía descripción
                df_evt["descripcion"] = df_evt.apply(
                    lambda r: desc_map.get(r["_k"], "") or r.get("descripcion", ""),
                    axis=1,
                )
                df_evt = df_evt.drop(columns="_k")
        except Exception:
            pass  # primera vez o no había hoja, sigue normal

    # ─── Preservar partidos creados a mano desde Streamlit ──────────────────
    # El usuario puede crear partidos vía la pestaña ✏️ Editar partido del
    # dashboard. Esos partidos NO existen en el Excel original. Cuando se
    # ejecuta el upload aquí, hay que conservar las filas con partido_ids
    # que NO están en el Excel para no destruir el trabajo manual.
    def _preservar_filas_manuales(hoja: str, df_excel: pd.DataFrame) -> pd.DataFrame:
        """Lee la hoja existente, separa las filas con partido_ids que NO
        están en df_excel (= datos creados manualmente) y las concatena
        con df_excel. Devuelve el DataFrame combinado."""
        if df_excel is None or df_excel.empty:
            return df_excel
        try:
            ws_old = sh.worksheet(hoja)
            old_rows = ws_old.get_all_records()
            old_df = pd.DataFrame(old_rows)
        except gspread.exceptions.WorksheetNotFound:
            return df_excel
        except Exception:
            return df_excel
        if old_df.empty or "partido_id" not in old_df.columns:
            return df_excel
        ids_excel = set(df_excel["partido_id"].astype(str).tolist())
        manuales = old_df[~old_df["partido_id"].astype(str).isin(ids_excel)].copy()
        if manuales.empty:
            return df_excel
        # Asegurar columnas comunes
        for c in df_excel.columns:
            if c not in manuales.columns:
                manuales[c] = ""
        for c in manuales.columns:
            if c not in df_excel.columns:
                df_excel[c] = ""
        # Mismo orden de columnas
        cols_orden = list(df_excel.columns)
        manuales = manuales[cols_orden]
        combinado = pd.concat([df_excel, manuales], ignore_index=True)
        print(f"  ℹ️  {hoja}: preservados {len(manuales)} filas de partidos "
              f"creados a mano ({manuales['partido_id'].nunique()} partidos).")
        return combinado

    print("Subiendo a Google Sheet:")
    df_jug_final = _preservar_filas_manuales("EST_PARTIDOS", df_jug)
    df_evt_final = _preservar_filas_manuales("EST_EVENTOS", df_evt)
    _write("EST_PARTIDOS", df_jug_final)
    _write("EST_EVENTOS", df_evt_final)
    _write("_VISTA_EST_JUGADOR", df_agr)
    if df_tot is not None and not df_tot.empty:
        df_tot_final = _preservar_filas_manuales("EST_TOTALES_PARTIDO", df_tot)
        _write("EST_TOTALES_PARTIDO", df_tot_final)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=XLSX_DEFAULT)
    ap.add_argument("--validar", action="store_true",
                    help="Imprime la validación contra HERRERO (cacheado en EST.TOTAL).")
    ap.add_argument("--upload", action="store_true",
                    help="Sube los datos al Google Sheet (hojas EST_PARTIDOS, EST_EVENTOS, _VISTA_EST_JUGADOR).")
    args = ap.parse_args()

    df_jug, df_evt, df_tot = procesar_excel(args.xlsx)

    print(f"Partidos procesados: {df_jug['partido_id'].nunique()}")
    print(f"Filas jugador-partido: {len(df_jug)}")
    print(f"Eventos de gol: {len(df_evt)}")
    print()

    # Resumen por jugador
    if not df_jug.empty:
        agr = df_jug.groupby("jugador", as_index=False).agg(
            partidos_conv=("convocado", "sum"),
            partidos_jug=("participa", "sum"),
            min_total=("min_total", "sum"),
            goles=("goles_a_favor", "sum"),
            asists=("asistencias", "sum"),
        ).sort_values("goles", ascending=False)
        print("Top 10 goleadores:")
        print(agr.head(10).to_string(index=False))
        print()

    if args.validar:
        print("─" * 60)
        print("VALIDACIÓN contra HERRERO (cacheado: 40 conv, 30 part, 1022 min)")
        h = df_jug[df_jug["jugador"] == "HERRERO"]
        if not h.empty:
            print(f"  Convocatorias calc:      {int(h['convocado'].sum())}")
            print(f"  Participaciones calc:    {int(h['participa'].sum())}")
            print(f"  Minutos calc:            {h['min_total'].sum():.1f}")
            print(f"  Goles a favor calc:      {int(h['goles_a_favor'].sum())}")
            print(f"  Asistencias calc:        {int(h['asistencias'].sum())}")

    if args.upload:
        df_agr = calcular_agregados_jugador(df_jug)
        subir_a_sheet(df_jug, df_evt, df_agr, df_tot)


if __name__ == "__main__":
    main()
