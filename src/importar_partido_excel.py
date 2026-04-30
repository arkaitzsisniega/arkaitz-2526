"""
importar_partido_excel.py — Importa UN partido desde el Excel
`Estadisticas2526.xlsx` a las hojas EST_* del Sheet maestro.

Modo preview (default): imprime lo que extraería SIN escribir.
Modo escribir: pasa --escribir para actualizar el Sheet.

Uso:
  /usr/bin/python3 src/importar_partido_excel.py "J27.PEÑISCOLA" \\
      --fecha 2026-04-29 --tipo LIGA

  Después de validar con preview:
      ... "J27.PEÑISCOLA" --fecha 2026-04-29 --escribir

Argumentos:
  hoja_excel       Nombre de la hoja del Excel (ej "J27.PEÑISCOLA")
  --fecha          ISO YYYY-MM-DD (no aparece en la hoja Excel)
  --tipo           LIGA | COPA DEL REY | COPA ESPAÑA | etc. (default LIGA)
  --escribir       Sin esto, solo preview. Con esto, escribe en el Sheet.
  --excel          Ruta al Excel (default: ~/Mi unidad/.../Estadisticas2526.xlsx)
"""
from __future__ import annotations

import argparse
import sys
import warnings
from datetime import time as dtime, timedelta
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parent.parent
EXCEL_DEFAULT = (Path.home() / "Mi unidad" / "Deporte" / "Futbol sala"
                  / "Movistar Inter" / "2025-26" / "Estadisticas"
                  / "Estadisticas2526.xlsx")
SHEET_NAME = "Arkaitz - Datos Temporada 2526"


# ─── Helpers ────────────────────────────────────────────────────────────────
def _td_a_minutos_decimal(v) -> float | None:
    """Excel guarda los tiempos como timedelta o float (días). Convertir
    a minutos decimal (ej 0:08:49 → 8.82)."""
    if v is None or v == "":
        return None
    if isinstance(v, timedelta):
        return round(v.total_seconds() / 60, 2)
    if isinstance(v, (int, float)):
        # Es float = días desde 1900-01-01. Para tiempos puros está entre 0-1.
        if 0 <= v < 2:
            return round(v * 24 * 60, 2)
        return float(v)
    if isinstance(v, str):
        # "0:20:00" formato H:M:S
        partes = v.strip().split(":")
        if len(partes) == 3:
            try:
                h, m, s = int(partes[0]), int(partes[1]), int(partes[2])
                return round(h * 60 + m + s / 60, 2)
            except ValueError:
                return None
    return None


def _td_a_seg(v) -> int | None:
    """Convertir timedelta/datetime.time/float/str a segundos enteros."""
    if v is None or v == "":
        return None
    if isinstance(v, timedelta):
        return int(v.total_seconds())
    if isinstance(v, dtime):
        # datetime.time(0, 1, 18) → 78 segundos
        return v.hour * 3600 + v.minute * 60 + v.second
    if isinstance(v, (int, float)):
        if 0 <= v < 2:
            return int(v * 24 * 3600)
        return int(v)
    if isinstance(v, str):
        partes = v.strip().split(":")
        if len(partes) == 3:
            try:
                return int(partes[0]) * 3600 + int(partes[1]) * 60 + int(partes[2])
            except ValueError:
                return None
    return None


def _seg_a_mmss(seg: int | None) -> str:
    if seg is None:
        return ""
    m, s = seg // 60, seg % 60
    return f"{m:02d}:{s:02d}"


def _to_int(v):
    if v is None or v == "":
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    try:
        return int(str(v).strip())
    except ValueError:
        return 0


# ─── Parsers ────────────────────────────────────────────────────────────────
def parsear_hoja(ws) -> dict:
    """Lee la hoja del Excel y devuelve un dict estructurado.

    Layout esperado (J27.PEÑISCOLA):
      Fila 3:  E="MOVISTAR INTER", I=rival, L="LIGA 25/26"
      Fila 5:  cabecera rotaciones (Nº, NOMBRE, 1ª Rot...8ª Rot, 1er Tiempo)
      Filas 6-19: Nº, Nombre, rotaciones (cols 4-11), col 12=1er Tiempo
      Filas 41-48: Resultado, MIN, ACCIÓN (goles)
      Filas 74-87: Stats T/S/NJ + minutos por parte
      Filas 91-104: Goles A FAVOR por jugador y tipo
      Filas 108-121: Goles EN CONTRA por jugador y tipo
      Filas 134-147: ESTADÍSTICAS INDIVIDUALES (MINS, PF, PNF, ROBOS, CORTES)
    """
    out = {
        "rival": "",
        "competicion": "",
        "local": "",
        "visitante": "",
        "marcador_final": "",
        "jugadores": [],     # [{dorsal, nombre, ts_njGOL, mins_total_seg, ...}]
        "goles": [],          # [{minuto_seg, accion_raw, marcador, equipo_marca}]
        "goles_a_favor_jug": {},   # nombre → {tipo: count}
        "goles_en_contra_jug": {}, # nombre → {tipo: count}
        "totales_equipo_af": {},   # tipo → count
        "totales_equipo_ec": {},
    }

    # Cabecera fila 3
    out["local"] = (ws.cell(3, 5).value or "").strip()
    out["visitante"] = (ws.cell(3, 9).value or "").strip()
    comp = (ws.cell(3, 12).value or "").strip()
    out["competicion"] = comp.replace(" 25/26", "").strip()
    # rival = el que NO sea Movistar Inter
    if "MOVISTAR" in out["local"].upper():
        out["rival"] = out["visitante"]
    else:
        out["rival"] = out["local"]

    # ── Plantilla con rotaciones (filas 6-19) ──────────────────────
    # Cols: B=Nº, C=Nombre, D-K=rotaciones (8), L=1er Tiempo
    # Pero hay también 2do tiempo. Voy a buscar columnas dinámicamente.
    # Revisando el Excel: las rotaciones de 2T están más a la derecha.
    fila_h = 5
    cabeceras_rot = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(fila_h, c).value
        cabeceras_rot.append((c, str(v).strip() if v else ""))

    # Localizar columnas relevantes
    col_num = None
    col_nombre = None
    cols_rot_1t = []  # 1ª Rot...8ª Rot del 1T
    col_total_1t = None
    cols_rot_2t = []
    col_total_2t = None
    col_total_match = None
    for c, txt in cabeceras_rot:
        txt_norm = txt.strip()
        txt_low = txt_norm.lower()
        if txt_norm == "Nº":
            col_num = c
        elif txt_norm == "NOMBRE":
            col_nombre = c
        elif "Rot" in txt_norm and "ª" in txt_norm:
            cols_rot_1t.append(c) if len(cols_rot_1t) < 8 else cols_rot_2t.append(c)
        elif txt_low in ("1er tiempo", "1º tiempo", "1ª parte"):
            col_total_1t = c
        elif txt_low in ("2o tiempo", "2º tiempo", "2do tiempo",
                          "2ª parte", "2o tiempo "):
            col_total_2t = c
        elif txt_low == "total":
            col_total_match = c

    # Si no hay 2do tiempo en las cabeceras, sigue (puede que solo tenga 1T)
    jugadores = []
    for r in range(6, 20):
        num = ws.cell(r, col_num).value if col_num else None
        nombre = ws.cell(r, col_nombre).value if col_nombre else None
        if not nombre:
            continue
        rot_1t = []
        for c in cols_rot_1t[:8]:
            seg = _td_a_seg(ws.cell(r, c).value)
            rot_1t.append(seg or 0)
        rot_2t = []
        for c in cols_rot_2t[:8]:
            seg = _td_a_seg(ws.cell(r, c).value)
            rot_2t.append(seg or 0)
        # Asegurar 8 elementos cada lista
        while len(rot_1t) < 8: rot_1t.append(0)
        while len(rot_2t) < 8: rot_2t.append(0)
        jugadores.append({
            "dorsal": _to_int(num),
            "nombre": str(nombre).strip().upper(),
            "rot_1t": rot_1t,
            "rot_2t": rot_2t,
            "min_1t_seg": _td_a_seg(ws.cell(r, col_total_1t).value)
                          if col_total_1t else None,
            "min_2t_seg": _td_a_seg(ws.cell(r, col_total_2t).value)
                          if col_total_2t else None,
        })
    out["jugadores"] = jugadores

    # ── Goles (filas 41-48) ────────────────────────────────────────
    for r in range(41, 50):
        marcador = ws.cell(r, 2).value
        minuto = ws.cell(r, 4).value
        accion = ws.cell(r, 6).value
        if not (marcador or accion):
            continue
        if str(marcador or "").strip().upper() == "RESULTADO":
            continue
        seg = _td_a_seg(minuto)
        accion_str = str(accion or "").strip().upper()
        equipo = ""
        accion_norm = accion_str
        if accion_str.startswith("AF."):
            equipo = "INTER"
            accion_norm = accion_str.replace("AF.", "").strip()
        elif accion_str.startswith("EC."):
            equipo = "RIVAL"
            accion_norm = accion_str.replace("EC.", "").strip()
        out["goles"].append({
            "minuto_seg": seg,
            "marcador": str(marcador).strip() if marcador else "",
            "accion_raw": accion_str,
            "accion": accion_norm,
            "equipo_marca": equipo,
        })

    # Marcador final = último gol
    if out["goles"]:
        out["marcador_final"] = out["goles"][-1]["marcador"]

    # ── Stats T/S/NJ (filas 74-87) ──────────────────────────────────
    # Cols: B=Nº, C=Nombre, D=T/S/NJ, F=GF, H=GC, J=DIF, L=Mins 1er Tiempo
    estados = {}  # nombre → {ts_nj, gf, gc, dif}
    for r in range(74, 88):
        num = ws.cell(r, 2).value
        nombre = ws.cell(r, 3).value
        ts_nj = ws.cell(r, 4).value
        gf = ws.cell(r, 6).value
        gc = ws.cell(r, 8).value
        if not nombre:
            continue
        nombre = str(nombre).strip().upper()
        estados[nombre] = {
            "ts_nj": str(ts_nj).strip().upper() if ts_nj else "",
            "gf": _to_int(gf),
            "gc": _to_int(gc),
        }
    # Mezclar al jugadores
    for j in out["jugadores"]:
        e = estados.get(j["nombre"])
        if e:
            j.update(e)

    # ── Goles a favor por jugador (filas 91-104) ───────────────────
    # Cols: C=Nº, D=Nombre, F=BANDA, G=CORNER, H=SAQUE CENTRO,
    #       I=FALTA, J=ABP 2ª JUGADA, K=10M, L=PENALTI
    tipos_gaf = ["BANDA", "CORNER", "SAQUE CENTRO", "FALTA", "ABP 2J", "10M", "PENALTI"]
    cols_gaf = [6, 7, 8, 9, 10, 11, 12]
    for r in range(91, 105):
        nombre = ws.cell(r, 4).value
        if not nombre:
            continue
        nombre = str(nombre).strip().upper()
        out["goles_a_favor_jug"][nombre] = {
            tipo: _to_int(ws.cell(r, c).value)
            for tipo, c in zip(tipos_gaf, cols_gaf)
        }

    # ── Goles en contra por jugador (filas 108-121) ────────────────
    for r in range(108, 122):
        nombre = ws.cell(r, 4).value
        if not nombre:
            continue
        nombre = str(nombre).strip().upper()
        out["goles_en_contra_jug"][nombre] = {
            tipo: _to_int(ws.cell(r, c).value)
            for tipo, c in zip(tipos_gaf, cols_gaf)
        }

    # ── Stats individuales (filas 134-147) ─────────────────────────
    # Cols: D=Nº, E=Jugador, G=MINS, H=PF, I=PNF, J=ROBOS, K=CORTES
    stats_ind = {}
    for r in range(134, 148):
        nombre = ws.cell(r, 5).value
        if not nombre:
            continue
        nombre = str(nombre).strip().upper()
        stats_ind[nombre] = {
            "pf": _to_int(ws.cell(r, 8).value),
            "pnf": _to_int(ws.cell(r, 9).value),
            "robos": _to_int(ws.cell(r, 10).value),
            "cortes": _to_int(ws.cell(r, 11).value),
        }
    for j in out["jugadores"]:
        s = stats_ind.get(j["nombre"])
        if s:
            j.update(s)

    return out


# ─── Preview ────────────────────────────────────────────────────────────────
def imprimir_preview(datos: dict, partido_id: str, fecha: str, tipo: str):
    print(f"\n{'='*70}")
    print(f"PARTIDO: {partido_id}")
    print(f"{'='*70}")
    print(f"  rival       : {datos['rival']}")
    print(f"  competición : {datos['competicion']}")
    print(f"  local/visit.: {datos['local']} vs {datos['visitante']}")
    print(f"  marcador    : {datos.get('marcador_final', '?')}")
    print(f"  fecha       : {fecha} (param)")
    print(f"  tipo        : {tipo}")

    print(f"\n--- Plantilla ({len(datos['jugadores'])} jugadores) ---")
    print(f"{'Nº':>3} | {'Nombre':<14} | {'T/S/NJ':<6} | {'Min1T':>6} | "
          f"{'Min2T':>6} | {'GF':>2} | {'GC':>2} | {'PF':>2} | {'PNF':>3}")
    for j in datos["jugadores"]:
        m1 = _seg_a_mmss(j.get("min_1t_seg"))
        m2 = _seg_a_mmss(j.get("min_2t_seg"))
        print(f"{j['dorsal']:>3} | {j['nombre']:<14} | "
              f"{j.get('ts_nj', ''):<6} | {m1:>6} | {m2:>6} | "
              f"{j.get('gf', 0):>2} | {j.get('gc', 0):>2} | "
              f"{j.get('pf', 0):>2} | {j.get('pnf', 0):>3}")

    print(f"\n--- Goles ({len(datos['goles'])}) ---")
    for g in datos["goles"]:
        print(f"  {_seg_a_mmss(g['minuto_seg'])} | {g['marcador']:<8} | "
              f"{g['equipo_marca']:<5} | {g['accion']}")


# ─── Escritura ──────────────────────────────────────────────────────────────
def escribir_a_sheet(datos: dict, partido_id: str, fecha: str, tipo: str):
    """Escribe en EST_PARTIDOS (1 fila por jugador) + EST_PLANTILLAS +
    EST_EVENTOS. EST_TOTALES_PARTIDO se rellena solo con cabecera (los
    detallados los meterá Arkaitz en el dashboard)."""
    import gspread
    from google.oauth2.service_account import Credentials
    SCOPES = ["https://www.googleapis.com/auth/spreadsheets",
              "https://www.googleapis.com/auth/drive"]
    creds = Credentials.from_service_account_file(
        str(ROOT / "google_credentials.json"), scopes=SCOPES)
    sh = gspread.authorize(creds).open(SHEET_NAME)

    rival = datos["rival"]
    competicion = datos["competicion"]
    local_es_inter = "MOVISTAR" in datos["local"].upper()
    local_visitante = "LOCAL" if local_es_inter else "VISITANTE"

    # ── EST_PARTIDOS: 1 fila por jugador ──
    ws_part = sh.worksheet("EST_PARTIDOS")
    headers_part = ws_part.row_values(1)
    filas_part = []
    for j in datos["jugadores"]:
        ts = j.get("ts_nj", "")
        convocado = ts in ("T", "S", "NJ")
        participa = ts in ("T", "S")
        rot_1t = j.get("rot_1t", [0]*8)
        rot_2t = j.get("rot_2t", [0]*8)
        # min_total = min_1t + min_2t en segundos → minutos:segundos float
        m1 = j.get("min_1t_seg") or 0
        m2 = j.get("min_2t_seg") or 0
        mtot = m1 + m2

        fila_dict = {
            "partido_id": partido_id,
            "tipo": tipo,
            "competicion": competicion,
            "rival": rival,
            "fecha": fecha,
            "dorsal": j["dorsal"],
            "jugador": j["nombre"],
            "min_1t": _seg_a_mmss(m1) if m1 else "",
            "min_2t": _seg_a_mmss(m2) if m2 else "",
            "min_total": _seg_a_mmss(mtot) if mtot else "",
            "convocado": "TRUE" if convocado else "FALSE",
            "participa": "TRUE" if participa else "FALSE",
            "pf": j.get("pf", 0),
            "pnf": j.get("pnf", 0),
            "robos": j.get("robos", 0),
            "cortes": j.get("cortes", 0),
            "goles_a_favor": j.get("gf", 0),
        }
        # Rotaciones
        for i, seg in enumerate(rot_1t, 1):
            fila_dict[f"rot_1t_{i}"] = _seg_a_mmss(seg) if seg else ""
        for i, seg in enumerate(rot_2t, 1):
            fila_dict[f"rot_2t_{i}"] = _seg_a_mmss(seg) if seg else ""
        # Construir fila en orden de headers
        fila = [str(fila_dict.get(h, "")) for h in headers_part]
        filas_part.append(fila)

    # ── EST_PLANTILLAS ──
    ws_plant = sh.worksheet("EST_PLANTILLAS")
    headers_plant = ws_plant.row_values(1)
    filas_plant = []
    for j in datos["jugadores"]:
        ts = j.get("ts_nj", "")
        if ts not in ("T", "S", "NJ"):
            continue
        # Posición: del Excel no la sabemos, miramos al roster por nombre
        # (lo deja en blanco; el dashboard ya lo cruza con JUGADORES_ROSTER)
        fila_dict = {
            "partido_id": partido_id,
            "tipo": tipo,
            "competicion": competicion,
            "rival": rival,
            "fecha": fecha,
            "dorsal": j["dorsal"],
            "jugador": j["nombre"],
            "posicion": "",
            "equipo": "INTER",
            "convocado": "TRUE",
        }
        filas_plant.append([str(fila_dict.get(h, "")) for h in headers_plant])

    # ── EST_EVENTOS ──
    ws_ev = sh.worksheet("EST_EVENTOS")
    headers_ev = ws_ev.row_values(1)
    filas_ev = []
    for g in datos["goles"]:
        seg = g.get("minuto_seg") or 0
        minuto_dec = round(seg / 60, 2) if seg else 0
        intervalo = ""
        if seg:
            mins = seg // 60
            ini = (mins // 5) * 5
            intervalo = f"{ini}-{ini+5}"
        fila_dict = {
            "partido_id": partido_id,
            "tipo": tipo,
            "competicion": competicion,
            "rival": rival,
            "fecha": fecha,
            "minuto": minuto_dec,
            "minuto_mmss": _seg_a_mmss(seg),
            "intervalo_5min": intervalo,
            "accion_raw": g["accion_raw"],
            "accion": g["accion"],
            "marcador": g["marcador"],
            "equipo_marca": g["equipo_marca"],
        }
        filas_ev.append([str(fila_dict.get(h, "")) for h in headers_ev])

    # ── EST_TOTALES_PARTIDO (solo cabecera + marcador) ──
    ws_tot = sh.worksheet("EST_TOTALES_PARTIDO")
    headers_tot = ws_tot.row_values(1)
    gf_total = sum(1 for g in datos["goles"] if g["equipo_marca"] == "INTER")
    gc_total = sum(1 for g in datos["goles"] if g["equipo_marca"] == "RIVAL")
    fila_tot_dict = {
        "partido_id": partido_id,
        "tipo": tipo,
        "competicion": competicion,
        "rival": rival,
        "fecha": fecha,
        "categoria": competicion,
        "lugar": "",
        "hora": "",
        "local_visitante": local_visitante,
        "goles_a_favor": gf_total,
        "goles_en_contra": gc_total,
    }
    fila_tot = [str(fila_tot_dict.get(h, "")) for h in headers_tot]

    # ── Comprobar si el partido ya existe → confirmar sobreescribir ──
    print("\n⚠️  Comprobando si el partido ya existe…")
    import time
    todos_part = ws_part.get_all_values()
    existe = any(
        len(r) > 0 and r[0].strip() == partido_id
        for r in todos_part[1:]
    )
    if existe:
        print(f"❗ El partido {partido_id} YA tiene filas en EST_PARTIDOS.")
        resp = input("¿Borrar y sobreescribir? [si/NO] ")
        if resp.strip().lower() not in ("si", "sí", "s", "yes", "y"):
            print("Abortado.")
            return
        # Borrar filas existentes
        for i in range(len(todos_part), 1, -1):
            if todos_part[i-1] and todos_part[i-1][0].strip() == partido_id:
                ws_part.delete_rows(i)
                time.sleep(0.4)
        # Lo mismo con plantillas, eventos, totales
        for ws_x in [ws_plant, ws_ev, ws_tot]:
            todos_x = ws_x.get_all_values()
            for i in range(len(todos_x), 1, -1):
                if todos_x[i-1] and todos_x[i-1][0].strip() == partido_id:
                    ws_x.delete_rows(i)
                    time.sleep(0.4)

    # ── Escribir todo ──
    print(f"\n📝 Escribiendo {len(filas_part)} filas en EST_PARTIDOS…")
    ws_part.append_rows(filas_part, value_input_option="USER_ENTERED")
    time.sleep(2)
    print(f"📝 Escribiendo {len(filas_plant)} filas en EST_PLANTILLAS…")
    ws_plant.append_rows(filas_plant, value_input_option="USER_ENTERED")
    time.sleep(2)
    print(f"📝 Escribiendo {len(filas_ev)} filas en EST_EVENTOS…")
    ws_ev.append_rows(filas_ev, value_input_option="USER_ENTERED")
    time.sleep(2)
    print(f"📝 Escribiendo cabecera en EST_TOTALES_PARTIDO…")
    ws_tot.append_row(fila_tot, value_input_option="USER_ENTERED")
    print("\n✅ Importación completa.")
    print(f"   Marcador final: {gf_total}-{gc_total}")
    print(f"   Convocados: {sum(1 for j in datos['jugadores'] if j.get('ts_nj') in ('T','S','NJ'))}")
    print(f"   Goles registrados: {len(datos['goles'])}")


# ─── Main ───────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                  formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("hoja_excel", help="Ej: J27.PEÑISCOLA")
    ap.add_argument("--fecha", required=True,
                     help="Fecha ISO YYYY-MM-DD (no aparece en la hoja)")
    ap.add_argument("--tipo", default="LIGA",
                     help="LIGA | COPA DEL REY | COPA ESPAÑA | etc.")
    ap.add_argument("--escribir", action="store_true",
                     help="Sin esto solo preview. Con esto escribe en el Sheet.")
    ap.add_argument("--excel", default=str(EXCEL_DEFAULT),
                     help=f"Ruta al Excel (default: {EXCEL_DEFAULT})")
    args = ap.parse_args()

    excel = Path(args.excel)
    if not excel.exists():
        print(f"❌ No encuentro el Excel: {excel}")
        return 1

    print(f"📂 Abriendo {excel.name}…")
    wb = openpyxl.load_workbook(str(excel), read_only=True, data_only=True)
    if args.hoja_excel not in wb.sheetnames:
        print(f"❌ La hoja '{args.hoja_excel}' no existe. Hojas disponibles:")
        for n in wb.sheetnames:
            print(f"   · {n}")
        return 1
    ws = wb[args.hoja_excel]
    datos = parsear_hoja(ws)

    partido_id = args.hoja_excel  # mismo formato que ya usa el sistema

    imprimir_preview(datos, partido_id, args.fecha, args.tipo)

    if args.escribir:
        escribir_a_sheet(datos, partido_id, args.fecha, args.tipo)
    else:
        print("\n💡 Esto era SOLO PREVIEW. Para escribir al Sheet, añade --escribir")
    return 0


if __name__ == "__main__":
    sys.exit(main())
