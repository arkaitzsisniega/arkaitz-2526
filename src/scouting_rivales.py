#!/usr/bin/env python3
"""
Importador de SCOUTING de rivales desde `Est. Goles rivales.xlsx`.

Origen: ~/Mi unidad/.../Estadisticas/Est. Goles rivales.xlsx
Hojas: 15 hojas con códigos de 3 letras (ALZ, BAR, CAR, ...).
Cada hoja: cómo marca ese rival cuando juega contra OTROS equipos
(no contra Inter). Datos partido a partido.

Estructura de cada hoja (filas 8+):
  Col B: Competición (ej. "LIGA 25/26")
  Col C: Equipo rival del scoutado (contra quién jugó)
  Col D: Fecha
  Col F: TOTAL goles a favor del scoutado
  Cols G-W: desglose por origen (BANDA, CÓRNER, SAQUE CENTRO, FALTA, ...)

Uso:
  /usr/bin/python3 src/scouting_rivales.py --upload
"""
from __future__ import annotations

import argparse
import datetime as _dt
import sys
import warnings
from pathlib import Path
from typing import Optional

warnings.filterwarnings("ignore")

import pandas as pd
def _col_letra(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


from openpyxl import load_workbook

XLSX_DEFAULT = (
    "/Users/mac/Mi unidad/Deporte/Futbol sala/Movistar Inter/"
    "2025-26/Estadisticas/Est. Goles rivales.xlsx"
)

# Diccionario código → nombre completo (temporada 25/26)
RIVALES_NOMBRES = {
    "ALZ": "Alzira FS",
    "BAR": "FC Barcelona",
    "CAR": "Jimbee Cartagena",
    "COR": "Córdoba Patrimonio",
    "ELP": "ElPozo Murcia",
    "IND": "Industrias Santa Coloma",
    "JAE": "Jaen Paraiso Interior",
    "MAN": "Manzanares Quesos Hidalgo",
    "NOI": "Noia Portus Apostoli",
    "OPA": "O Parrulo",
    "PAL": "Palma Futsal",
    "PEÑ": "Peñiscola Rehabmedic",
    "RIB": "Ribera de Navarra",
    "VAL": "Valdepeñas Viña Albali",
    "XOT": "Osasuna Magna",
}

# Bloque 1: cols 6-27 = TOTAL A/F + 21 acciones de goles A FAVOR
# Bloque 1: cols 29-50 = TOTAL E/C + 21 acciones de goles EN CONTRA
# Bloque 2: cols 55-94 = zonas (portería P1-P9 y campo Z1-Z11) AF y EC

# Columnas A FAVOR (origen del gol)
COLS_AF_ORIGEN = {
    7:  "Banda", 8:  "Córner", 9:  "Saque de Centro", 10: "Falta",
    11: "2ª jugada", 12: "10 metros", 13: "Penalti", 14: "Falta sin barrera",
    15: "Salida de presión", 16: "Ataque Posicional 4x4", 17: "1x1 en banda",
    18: "2ª jugada de ABP", 19: "Incorporación del portero",
    20: "Robo en incorporación de portero",
    21: "5x4", 22: "4x3", 23: "4x5", 24: "3x4",
    25: "Contraataque", 26: "Robo en zona alta", 27: "No calificado",
}
# Columnas EN CONTRA (origen del gol que les meten)
COLS_EC_ORIGEN = {
    30: "Banda", 31: "Córner", 32: "Saque de Centro", 33: "Falta",
    34: "2ª jugada", 35: "10 metros", 36: "Penalti", 37: "Falta sin barrera",
    38: "Salida de presión", 39: "Ataque Posicional 4x4", 40: "1x1 en banda",
    41: "2ª jugada de ABP", 42: "Incorporación del portero",
    43: "Pérdida en incorporación de portero",
    44: "5x4", 45: "4x3", 46: "4x5", 47: "3x4",
    48: "Contraataque", 49: "Robo en zona alta", 50: "No calificado",
}
# Zonas A FAVOR: 9 cuadrantes de portería (cols 55-63) + 11 zonas campo (cols 64-74)
COLS_AF_PORT = {55+i: f"P{i+1}" for i in range(9)}
COLS_AF_ZONA = {64+i: f"Z{i+1}" for i in range(11)}
# Zonas EN CONTRA: 9 cuadrantes portería (cols 75-83) + 11 zonas campo (cols 84-94)
# (en el Excel las cabeceras de cols 75-83 están como "G.AF P" pero son
#  realmente las zonas de portería de los goles EN CONTRA, según contexto)
COLS_EC_PORT = {75+i: f"P{i+1}" for i in range(9)}
COLS_EC_ZONA = {84+i: f"Z{i+1}" for i in range(11)}


def _to_int(v) -> int:
    if v is None or v == "":
        return 0
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except Exception:
            return 0


def _to_date_iso(v) -> str:
    if isinstance(v, _dt.datetime):
        return v.date().isoformat()
    if isinstance(v, _dt.date):
        return v.isoformat()
    return ""


def cargar(xlsx_path: str = XLSX_DEFAULT) -> pd.DataFrame:
    wb = load_workbook(xlsx_path, data_only=True)
    filas: list[dict] = []

    for codigo, nombre_full in RIVALES_NOMBRES.items():
        if codigo not in wb.sheetnames:
            print(f"⚠️  Hoja '{codigo}' no encontrada, salto.")
            continue
        ws = wb[codigo]
        # Datos a partir de fila 8 (1-indexed)
        for row in ws.iter_rows(min_row=8, values_only=True):
            if not row:
                continue
            # Col idx 1-based en openpyxl ↔ 0-based en tuple
            def _g(c1):
                return row[c1 - 1] if len(row) >= c1 else None

            comp = _g(2)
            contra = _g(3)
            fecha = _g(4)
            if not comp or not contra:
                continue
            base = {
                "rival_codigo": codigo,
                "rival_nombre": nombre_full,
                "competicion": str(comp).strip(),
                "contra_quien": str(contra).strip(),
                "fecha": _to_date_iso(fecha),
                "total_a_favor": _to_int(_g(6)),
                "total_en_contra": _to_int(_g(29)),
            }
            # Conteos por origen — A FAVOR (cols 7-27)
            for col_idx, accion in COLS_AF_ORIGEN.items():
                base[f"AF_{accion}"] = _to_int(_g(col_idx))
            # Conteos por origen — EN CONTRA (cols 30-50)
            for col_idx, accion in COLS_EC_ORIGEN.items():
                base[f"EC_{accion}"] = _to_int(_g(col_idx))
            # Zonas portería A FAVOR (cols 55-63: P1-P9)
            for col_idx, p in COLS_AF_PORT.items():
                base[f"AF_port_{p}"] = _to_int(_g(col_idx))
            # Zonas campo A FAVOR (cols 64-74: Z1-Z11)
            for col_idx, z in COLS_AF_ZONA.items():
                base[f"AF_zona_{z}"] = _to_int(_g(col_idx))
            # Zonas portería EN CONTRA (cols 75-83: P1-P9)
            for col_idx, p in COLS_EC_PORT.items():
                base[f"EC_port_{p}"] = _to_int(_g(col_idx))
            # Zonas campo EN CONTRA (cols 84-94: Z1-Z11)
            for col_idx, z in COLS_EC_ZONA.items():
                base[f"EC_zona_{z}"] = _to_int(_g(col_idx))
            filas.append(base)

    return pd.DataFrame(filas)


def calcular_agregado_rival(df: pd.DataFrame) -> pd.DataFrame:
    """Una fila por rival con totales y % por origen.

    Incluye AF (cómo marca), EC (cómo recibe) y zonas (P y Z) en ambos.
    """
    if df.empty:
        return pd.DataFrame()
    cols_sum = [c for c in df.columns
                if c.startswith(("AF_", "EC_", "total_"))]
    aggs = {"partidos": ("contra_quien", "count")}
    for c in cols_sum:
        aggs[c] = (c, "sum")
    g = df.groupby(["rival_codigo", "rival_nombre"], as_index=False).agg(**aggs)

    # % por origen sobre total_a_favor
    for accion in COLS_AF_ORIGEN.values():
        col = f"AF_{accion}"
        if col in g.columns:
            g[f"%AF_{accion}"] = (
                g[col] / g["total_a_favor"].replace(0, float("nan")) * 100
            ).round(1)
    # % por origen sobre total_en_contra
    for accion in COLS_EC_ORIGEN.values():
        col = f"EC_{accion}"
        if col in g.columns:
            g[f"%EC_{accion}"] = (
                g[col] / g["total_en_contra"].replace(0, float("nan")) * 100
            ).round(1)
    return g.sort_values("total_a_favor", ascending=False)


def subir_a_sheet(df_raw: pd.DataFrame, df_agr: pd.DataFrame) -> None:
    import gspread
    from google.oauth2.service_account import Credentials
    SCOPES = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file("google_credentials.json", scopes=SCOPES)
    gc = gspread.authorize(creds)
    sh = gc.open("Arkaitz - Datos Temporada 2526")

    def _write(hoja: str, df: pd.DataFrame):
        try:
            ws = sh.worksheet(hoja)
            ws.clear()
        except gspread.exceptions.WorksheetNotFound:
            ws = sh.add_worksheet(title=hoja, rows=max(len(df) + 5, 30), cols=max(len(df.columns), 6))
        out = df.where(pd.notnull(df), "")
        valores = [list(out.columns)] + out.astype(str).values.tolist()
        ws.update(values=valores, range_name="A1")
        ws.format(f"A1:{_col_letra(len(out.columns))}1", {"textFormat": {"bold": True}})
        print(f"✅ {hoja}: {len(out)} filas, {len(out.columns)} cols")

    _write("SCOUTING_RIVALES", df_raw)
    _write("_VISTA_SCOUTING_RIVAL", df_agr)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=XLSX_DEFAULT)
    ap.add_argument("--upload", action="store_true")
    args = ap.parse_args()

    df_raw = cargar(args.xlsx)
    df_agr = calcular_agregado_rival(df_raw)

    print(f"Filas raw: {len(df_raw)}")
    print(f"Rivales scouted: {df_raw['rival_codigo'].nunique()}")
    print()
    print("Top rivales por goles registrados:")
    print(df_agr[["rival_codigo", "rival_nombre", "partidos", "total_a_favor", "total_en_contra"]]
          .head(15).to_string(index=False))

    if args.upload:
        subir_a_sheet(df_raw, df_agr)


if __name__ == "__main__":
    main()
